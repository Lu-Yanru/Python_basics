#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automatic process excel spreadsheets

Created on Tue Dec 31 15:47:36 2024

@author: luzi
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


# load an excel workbook
wb = xl.load_workbook("transactions.xlsx")
# access excel sheet
sheet = wb["Sheet1"]
# access a cell
cell = sheet["a1"]
# or through coordinate (row, col)
cell = sheet.cell(1, 1)
print(cell.value)
# how many rows are in the sheet
print(sheet.max_row)

# plus 1 because range does not include the last value
# also excludes the title row
for row in range(2, sheet.max_row + 1):
    # take the cells in the 3rd column
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    # write the corrected value in a new column
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    
# add a chart to the current sheet
# select a range of values
values = Reference(sheet, 
          min_row=2, 
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
# create a chart
chart = BarChart()
# add values to the chart
chart.add_data(values)
# add chart to sheet and where (cell coordinate)
sheet.add_chart(chart, "e2")

wb.save("transactions2.xlsx")
 

# define the whole thing as a function to automate
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    # access excel sheet
    sheet = wb["Sheet1"]

    # plus 1 because range does not include the last value
    # also excludes the title row
    for row in range(2, sheet.max_row + 1):
        # take the cells in the 3rd column
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        # write the corrected value in a new column
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        
    # add a chart to the current sheet
    # select a range of values
    values = Reference(sheet, 
              min_row=2, 
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    # create a chart
    chart = BarChart()
    # add values to the chart
    chart.add_data(values)
    # add chart to sheet and where (cell coordinate)
    sheet.add_chart(chart, "e2")

    wb.save(filename)
    
# loop through all excel files in the directory and automatically process all files
path = Path()

for file in path.glob("*.xlsx"):
    process_workbook(file)